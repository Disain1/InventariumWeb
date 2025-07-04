from flask import Flask, render_template, request, jsonify, send_file
import firebase_admin
from firebase_admin import credentials, firestore, storage
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as ExcelImage

import qrcode
import pandas as pd
import io
import uuid

app = Flask(__name__)

# Инициализация Firebase
cred = credentials.Certificate('creds.json')
firebase_admin.initialize_app(cred, {
    'storageBucket': 'test-storage-73eb7.firebasestorage.app'
})
bucket = storage.bucket()
db = firestore.client()

items_ref = db.collection('items')
storages_ref = db.collection('storages')


def clean_value(val):
    return '' if pd.isna(val) else str(val).strip()


def clean_value_int(val):
    return 1 if pd.isna(val) else int(str(val).strip())


def get_doc_data(doc):
    if doc != None:
        data = doc.to_dict() if doc.exists else {}
        # Убираем поля, которые часто меняются автоматически
        data.pop('recentChangeTimestamp', None)
        return data


def upload_file_to_storage(file):
    filename = 'images/' + str(uuid.uuid4()) + '_' + file.filename
    blob = bucket.blob(filename)
    blob.upload_from_file(file, content_type=file.content_type)
    blob.make_public()  # Чтобы получить публичную ссылку
    return blob.public_url


@app.route('/check_storage_items_count/<storage_id>')
def check_storage_items_count(storage_id):
    storage_ref = storages_ref.document(storage_id)
    items_query = items_ref.where('storage', '==', storage_ref).stream()
    count = len(list(item.to_dict() for item in items_query))
    print(count)
    return jsonify({'count': count})


@app.route('/add_storage', methods=['POST'])
def add_storage():
    name = request.form.get('name')
    if not name:
        return jsonify({'success': False}), 400

    photo_url = request.form.get('photoUrl', '')
    if 'photoFile' in request.files and request.files['photoFile'].filename:
        photo_url = upload_file_to_storage(request.files['photoFile'])

    storage_data = {
        'name': name,
        'note': request.form.get('note', ''),
        'address': request.form.get('address', ''),
        'recentChangeTimestamp': firestore.SERVER_TIMESTAMP,
        'recentChangeUser': 'admin',
        'photoUrl': photo_url
    }

    new_doc = storages_ref.document()
    storage_data['id'] = new_doc.id
    new_doc.set(storage_data)

    return jsonify({'success': True})



@app.route('/add_item', methods=['POST'])
def add_item():
    name = request.form.get('name')
    storage_id = request.form.get('storage')
    if not (name and storage_id):
        return jsonify({'success': False}), 400

    photo_url = request.form.get('photoUrl', '')
    if 'photoFile' in request.files and request.files['photoFile'].filename:
        photo_url = upload_file_to_storage(request.files['photoFile'])

    item_data = {
        'article': request.form.get('article', ''),
        'name': name,
        'count': int(request.form.get('count', 1)),
        'note': request.form.get('note', ''),
        'location': request.form.get('location', ''),
        'recentChangeUser': 'admin',
        'recentChangeTimestamp': firestore.SERVER_TIMESTAMP,
        'photoUrl': photo_url,
        'storage': storages_ref.document(storage_id)
    }

    new_doc = items_ref.document()
    item_data['id'] = new_doc.id
    new_doc.set(item_data)

    return jsonify({'success': True})


@app.route('/update_item', methods=['POST'])
def update_item():
    doc_id = request.form.get('doc_id')
    if not doc_id:
        return jsonify({'success': False, 'message': 'Не передан ID документа'}), 400

    doc_ref = items_ref.document(doc_id)
    doc_snapshot = doc_ref.get()
    if not doc_snapshot.exists:
        return jsonify({'success': False, 'message': 'Документ не найден'}), 404

    update_data = {}

    def get_clean(key, default=''):
        return request.form.get(key, default).strip()

    update_data['name'] = get_clean('name')
    update_data['article'] = get_clean('article')
    update_data['note'] = get_clean('note')
    update_data['location'] = get_clean('location')

    # Количество
    try:
        update_data['count'] = int(request.form.get('count', 1))
    except ValueError:
        update_data['count'] = 1

    # Склад
    storage_id = get_clean('storage')
    if storage_id:
        update_data['storage'] = storages_ref.document(storage_id)

    # Фото: либо URL, либо новый файл
    photo_url = get_clean('photoUrl')
    photo_file = request.files.get('photoFile')

    if photo_file and photo_file.filename:
        # Загружаем фото в хранилище
        photo_url = upload_file_to_storage(photo_file)

    if photo_url:
        update_data['photoUrl'] = photo_url

    update_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP
    update_data['recentChangeUser'] = 'admin'

    doc_ref.update(update_data)
    return jsonify({'success': True})


@app.route('/import_excel', methods=['POST'])
def import_excel():
    file = request.files.get('file')
    if not file:
        return "Нет файла", 400

    df_storages = pd.read_excel(file, sheet_name='Storages')
    df_items = pd.read_excel(file, sheet_name='Items')

    # === Получаем текущие ID из Firestore ===
    current_storage_docs = {doc.id: doc for doc in storages_ref.stream()}
    current_item_docs = {doc.id: doc for doc in items_ref.stream()}

    excel_storage_ids = set()
    excel_item_ids = set()
    storage_name_to_ref = {}

    # Для лога
    log = {
        'added_storages': [],
        'updated_storages': [],
        'deleted_storages': [],
        'added_items': [],
        'updated_items': [],
        'deleted_items': []
    }

    # === Импорт/обновление складов ===
    for _, row in df_storages.iterrows():
        storage_id = str(row.get('ID склада', '')).strip()
        if not storage_id:
            continue  # пропускаем без ID
        excel_storage_ids.add(storage_id)

        doc_ref = storages_ref.document(storage_id)
        old_data = get_doc_data(current_storage_docs.get(storage_id, None))

        doc_snapshot = doc_ref.get()

        storage_data = {
            'address': clean_value(row.get('Адрес', '')),
            'id': '',
            'photoUrl': clean_value(row.get('Ссылка на фото')),
            'name': clean_value(row.get('Наименование', '')),
            'note': clean_value(row.get('Примечание', '')),
            'recentChangeUser': row.get('Логин', ''),
            #'recentChangeTimestamp': firestore.SERVER_TIMESTAMP
        }

        if not doc_snapshot.exists:
            storage_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP

        if not old_data:
            log['added_storages'].append(storage_data['name'])
        elif any(storage_data.get(k) != old_data.get(k) for k in ['address', 'photoUrl', 'name', 'note']):
            log['updated_storages'].append(storage_data['name'])



        doc_ref.set(storage_data, merge=True)
        storage_name_to_ref[row.get('Наименование', '')] = doc_ref

    # === Удаление складов, которых нет в Excel ===
    for firestore_id in current_storage_docs:
        if firestore_id not in excel_storage_ids:
            name = current_storage_docs[firestore_id].to_dict().get("name", firestore_id)
            log['deleted_storages'].append(name)

            storages_ref.document(firestore_id).delete()

    # === Импорт/обновление вещей ===
    for _, row in df_items.iterrows():
        item_id = str(row.get('ID вещи', '')).strip()
        if not item_id:
            continue  # пропускаем без ID
        excel_item_ids.add(item_id)

        storage_name = row.get('Склад')
        storage_ref = storage_name_to_ref.get(storage_name)
        if not storage_ref:
            continue  # склад не найден

        doc_ref = items_ref.document(item_id)
        old_data = get_doc_data(current_item_docs.get(item_id, None))

        doc_snapshot = doc_ref.get()

        item_data = {
            'article': clean_value(row.get('Артикул', '')),
            'id': '',
            'name': clean_value(row.get('Наименование', '')),
            'note': clean_value(row.get('Примечание', '')),
            'count': clean_value_int(row.get("Количество", '')),
            'recentChangeUser': clean_value(row.get('Логин', '')),
            #'recentChangeTimestamp': firestore.SERVER_TIMESTAMP,
            'location': clean_value(row.get('Расположение на складе', '')),
            'photoUrl': clean_value(row.get('Ссылка на фото', '')),
            'storage': storage_ref
        }

        if not doc_snapshot.exists:
            item_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP

        if not old_data:
            log['added_items'].append(item_data['name'])
        elif any(item_data.get(k) != old_data.get(k) for k in ['article', 'name', 'note', 'count', 'location']):
            log['updated_items'].append(item_data['name'])

        doc_ref.set(item_data, merge=True)

    # === Удаление вещей, которых нет в Excel ===
    for firestore_id in current_item_docs:
        if firestore_id not in excel_item_ids:
            name = current_item_docs[firestore_id].to_dict().get("name", firestore_id)
            log['deleted_items'].append(name)
            items_ref.document(firestore_id).delete()


    # === Формируем HTML лог ===
    html_log = "<h2>Импорт завершён успешно!</h2>"
    for category, items in log.items():
        if not items:
            continue
        action = {
            'added_storages': 'Добавлены склады',
            'updated_storages': 'Обновлены склады',
            'deleted_storages': 'Удалены склады',
            'added_items': 'Добавлены вещи',
            'updated_items': 'Обновлены вещи',
            'deleted_items': 'Удалены вещи'
        }.get(category, category)
        html_log += f"<h3>{action}:</h3><ul>"
        for name in items:
            html_log += f"<li>{name}</li>"
        html_log += "</ul>"

    return html_log, 200


@app.route('/export_excel')
def export_excel():
    # Получаем данные о складах
    storages = {doc.id: doc.to_dict()['name'] for doc in storages_ref.stream()}

    # Получаем данные о товарах
    items_data = []
    for doc in items_ref.stream():
        item = doc.to_dict()

        # Добавляем название склада, заменяя ссылку на строку
        storage_name = storages.get(item['storage'].id, 'Не найдено')
        items_data.append({
            'ID вещи': doc.id,
            'Артикул': item.get('article'),
            'Наименование': item.get('name'),
            'Количество': item.get('count'),
            'Примечание': item.get('note'),
            'Логин': item.get('recentChangeUser'),
            'Расположение на складе': item.get('location'),
            'Склад': storage_name,
            'Ссылка на фото': item.get('photoUrl')
        })

    # Создаём DataFrame для товаров (items)
    items_df = pd.DataFrame(items_data)

    # Создаём рабочую книгу Excel
    wb = Workbook()

    # Лист для Items
    items_sheet = wb.active
    items_sheet.title = 'Items'

    # Заголовки
    # headers = ['ID вещи', 'Артикул', 'Наименование', 'Количество', 'Примечание', 'Логин', 'Расположение на складе', 'Склад', "Ссылка на фото"]
    # items_sheet.append(headers)

    headers = ['ID вещи', 'Артикул', 'Наименование', 'Количество', 'Примечание', 'Логин', 'Расположение на складе', 'Склад', "Ссылка на фото", 'QR-код']
    items_sheet.append(headers)
    

    # Форматирование для заголовков
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, cell in enumerate(items_sheet[1], 1):
        cell.fill = header_fill
        cell.font = header_font

    # Задаём ширину колонок для Items
    items_sheet.column_dimensions['A'].width = 30  # ID вещи
    items_sheet.column_dimensions['B'].width = 20  # Артикул
    items_sheet.column_dimensions['C'].width = 30  # Наименование
    items_sheet.column_dimensions['D'].width = 10   # Количество
    items_sheet.column_dimensions['E'].width = 40  # Примечание
    items_sheet.column_dimensions['F'].width = 35  # Логин
    items_sheet.column_dimensions['G'].width = 30  # Расположение на складе
    items_sheet.column_dimensions['I'].width = 30  # Склад (название)
    items_sheet.column_dimensions['J'].width = 30  # QR CODE


    # # Добавляем строки данных
    # for row in items_data:
    #     items_sheet.append([row['ID вещи'], row['Артикул'], row['Наименование'], row['Количество'], row['Примечание'], row['Логин'], row['Расположение на складе'], row['Склад'], row['Ссылка на фото']])

    for i, row in enumerate(items_data, start=2):  # начинаем со 2-й строки (после заголовка)
        items_sheet.row_dimensions[i].height = 64

        items_sheet.append([
            row['ID вещи'], row['Артикул'], row['Наименование'], row['Количество'],
            row['Примечание'], row['Логин'], row['Расположение на складе'],
            row['Склад'], row['Ссылка на фото']
        ])

        # Генерируем QR-код
        qr_text = f"items/{row['ID вещи']}"
        qr_img = qrcode.make(qr_text, box_size=20)

        img_path = f"tmp/item_qr_{i}.png"
        qr_img.save(img_path)

        img = ExcelImage(img_path)
        img.width = 64
        img.height = 64
        cell_ref = f"J{i}"  # 10-я колонка
        items_sheet.add_image(img, cell_ref)

    # Лист для Storages
    storages_sheet = wb.create_sheet(title="Storages")

    # Заголовки для складов
    storages_headers = ['ID склада', 'Наименование', 'Примечание', 'Адрес', 'Логин', "Ссылка на фото", "QR-код"]
    storages_sheet.append(storages_headers)

    # Форматирование заголовков
    for cell in storages_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    storages_sheet.column_dimensions['A'].width = 20  # ID склада
    storages_sheet.column_dimensions['B'].width = 30  # Наименование
    storages_sheet.column_dimensions['C'].width = 40  # Примечание
    storages_sheet.column_dimensions['D'].width = 40  # Примечание
    storages_sheet.column_dimensions['E'].width = 35  # Последнее изменение
    storages_sheet.column_dimensions['G'].width = 15  # QR CODE

    # Добавляем данные складов
    for i, doc in enumerate(storages_ref.stream(), start=2):
        storages_sheet.row_dimensions[i].height = 64
        storage = doc.to_dict()
        storages_sheet.append([
            doc.id,
            storage.get('name', ''),
            storage.get('note', ''),
            storage.get('address', ''),
            storage.get('recentChangeUser', ''),
            storage.get('photoUrl')
        ])

        qr_text = f"storages/{doc.id}"
        qr_img = qrcode.make(qr_text)
        img_path = f"tmp/storage_qr_{i}.png"
        qr_img.save(img_path)

        img = ExcelImage(img_path)
        img.width = 64
        img.height = 64
        cell_ref = f"G{i}"  # 7-я колонка
        storages_sheet.add_image(img, cell_ref)

    # Сохраняем файл в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="inventory_export.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/')
def index():
    storages = sorted(
        [doc.to_dict() | {'id': doc.id} for doc in storages_ref.stream()],
        key=lambda x: x.get('name', '').lower()
    )

    items = []
    for doc in items_ref.stream():
        data = doc.to_dict()
        data['doc_id'] = doc.id
        items.append(data)

    return render_template('index.html', items=items, storages=storages)


@app.route('/move_item', methods=['POST'])
def move_item():
    data = request.get_json()
    doc_id = data.get('doc_id')
    move_count = int(data.get('count', 0))
    destination_id = data.get('destination')

    if not doc_id or not destination_id or move_count <= 0:
        return jsonify({'success': False}), 400

    item_doc = items_ref.document(doc_id).get()
    if not item_doc.exists:
        return jsonify({'success': False, 'message': 'Вещь не найдена'}), 404

    item = item_doc.to_dict()
    if item['count'] < move_count:
        return jsonify({'success': False, 'message': 'Недостаточно количества'}), 400

    name = item['name']
    destination_ref = storages_ref.document(destination_id)

    # Проверяем, есть ли такая же вещь на целевом складе
    existing_q = items_ref.where('name', '==', name).where('storage', '==', destination_ref).stream()
    existing_item = next(existing_q, None)

    if existing_item:
        existing_data = existing_item.to_dict()
        new_count = existing_data.get('count', 1) + move_count
        items_ref.document(existing_item.id).update({
            'count': new_count,
            'recentChangeTimestamp': firestore.SERVER_TIMESTAMP
        })
        message = 'Вещь объединена с уже существующей.'
    else:
        # Копируем данные с новой ссылкой и новым количеством
        new_data = item.copy()
        new_data['count'] = move_count
        new_data['storage'] = destination_ref
        new_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP
        new_data['recentChangeUser'] = 'admin'
        new_doc = items_ref.document()
        new_data['id'] = new_doc.id
        new_doc.set(new_data)
        message = 'Вещь перемещена на новый склад.'

    # Уменьшаем количество или удаляем исходную
    if item['count'] == move_count:
        items_ref.document(doc_id).delete()
    else:
        items_ref.document(doc_id).update({
            'count': item['count'] - move_count,
            'recentChangeTimestamp': firestore.SERVER_TIMESTAMP
        })

    return jsonify({'success': True, 'message': message})


@app.route('/update_cell', methods=['POST'])
def update_cell():
    data = request.json
    doc_id = data.get('doc_id')
    field = data.get('field')
    value = data.get('value')
    collection = data.get('collection', 'items')  # default to 'items'

    if not (doc_id and field and collection):
        return jsonify({'success': False}), 400

    update_data = {field: value}
    doc_ref = db.collection(collection).document(doc_id)

    if field == 'storage':
        # Получаем текущую вещь
        current_doc = doc_ref.get()
        if not current_doc.exists:
            return jsonify({'success': False}), 404
        current_item = current_doc.to_dict()
        current_name = current_item.get('name')
        current_count = current_item.get('count', 1)
        new_storage_ref = storages_ref.document(value)

        # Ищем вещь с таким же названием на новом складе
        same_name_query = items_ref.where('name', '==', current_name).where('storage', '==', new_storage_ref).stream()
        matching_item = next(same_name_query, None)

        if matching_item:
            existing_item = matching_item.to_dict()
            existing_doc_id = matching_item.id
            total_count = existing_item.get('count', 1) + current_count

            # Обновляем количество на новом складе
            items_ref.document(existing_doc_id).update({
                'count': total_count,
                'recentChangeTimestamp': firestore.SERVER_TIMESTAMP
            })

            # Удаляем текущую вещь
            doc_ref.delete()

            return jsonify({'success': True, 'merged': True})

        else:
            # Просто обновляем склад
            update_data['storage'] = new_storage_ref
            update_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP
            doc_ref.update(update_data)
            return jsonify({'success': True, 'moved': True})

    if field == 'count':
        update_data[field] = int(value) if str(value).isnumeric() else 1

    update_data['recentChangeTimestamp'] = firestore.SERVER_TIMESTAMP
    update_data['recentChangeUser'] = 'admin'
    doc_ref.update(update_data)
    return jsonify({'success': True})


# @app.route('/delete_document', methods=['POST'])
# def delete_document():
#     data = request.json
#     doc_id = data.get('doc_id')
#     collection = data.get('collection', 'items')

#     if not doc_id or collection not in ['items', 'storages']:
#         return jsonify({'success': False}), 400
    
#     if collection == 'storages':
#         db.collection(collection)

#     db.collection(collection).document(doc_id).delete()
#     return jsonify({'success': True})


@app.route('/delete_document', methods=['POST'])
def delete_document():
    data = request.json
    doc_id = data.get('doc_id')
    collection = data.get('collection', 'items')

    if not doc_id or collection not in ['items', 'storages']:
        return jsonify({'success': False}), 400

    if collection == 'storages':
        # Удаляем все вещи, у которых storage == ссылка на этот склад
        storage_ref = storages_ref.document(doc_id)
        items_to_delete = items_ref.where('storage', '==', storage_ref).stream()
        deleted_names = []

        for item_doc in items_to_delete:
            items_ref.document(item_doc.id).delete()
            deleted_names.append(item_doc.to_dict().get('name', item_doc.id))

        # Удаляем сам склад
        storages_ref.document(doc_id).delete()

        return jsonify({'success': True, 'deleted_items': deleted_names})

    # Если просто удаляется вещь
    db.collection(collection).document(doc_id).delete()
    return jsonify({'success': True})


if __name__ == "__main__":
   app.run(debug=True)